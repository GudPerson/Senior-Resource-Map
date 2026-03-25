#!/usr/bin/env python3
from __future__ import annotations

import concurrent.futures
import copy
import datetime as dt
import html as html_lib
import json
import random
import re
import time
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlencode, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


ROOT = Path("/Users/sweetbuns/Documents/New project")
TMP_DIR = ROOT / "tmp" / "spreadsheets"
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
TEMPLATE_PATH = Path("/Users/sweetbuns/Downloads/Community resource Template.xlsx")
POSTAL_CODES_PATH = Path("/Users/sweetbuns/Downloads/SG Postal codes.xlsx")
OUTPUT_PATH = OUTPUT_DIR / "singapore_community_resources_best_effort.xlsx"

TODAY = dt.date(2026, 3, 25)
HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    )
}

TEMPLATE_HEADERS = [
    "[DROPDOWN] Category",
    "[DROPDOWN] Type",
    "[FREE TEXT] Name",
    "[FREE TEXT] Description\n(incl. access to which Service or Programme)",
    "[DROPDOWN] Cost",
    "[FREE TEXT] Service Provider",
    "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100",
    "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri",
    "[FREE TEXT] Address",
    "[FREE TEXT] Postal Code",
    "[FREE TEXT] Contact Number",
    "[FREE TEXT] Email",
    "[FREE TEXT] Website/URL",
    "Lat",
    "Long",
    "[DROPDOWN] Status",
    "Planning Area",
    "Subzone",
    "Constituency",
    "RHS",
    "Region",
    "[DROPDOWN] Sub-region",
]

AIC_PAGES: dict[str, dict[str, str]] = {
    "aac": {
        "url": "https://aic.sg/Care-Services/Active-Ageing-Centres",
        "category": "Facility (AIC-related)",
        "type": "AAC",
        "description": "Active ageing programmes and social activities for seniors.",
        "provider": "AIC",
    },
    "day_care": {
        "url": "https://aic.sg/Care-Services/Day-Care",
        "category": "Facility (AIC-related)",
        "type": "SCC",
        "description": "Senior day care support for older persons who need supervision.",
        "provider": "AIC",
    },
    "nursing_home": {
        "url": "https://aic.sg/Care-Services/Nursing-Home",
        "category": "Facility (AIC-related)",
        "type": "Nursing Home",
        "description": "Residential nursing home care for seniors with higher care needs.",
        "provider": "AIC",
    },
    "aic_link": {
        "url": "https://aic.sg/Contact-Us",
        "category": "Facility (AIC-related)",
        "type": "AIC LINK",
        "description": "Advice on care services, schemes, caregiver support and ageing-in-place options.",
        "provider": "AIC",
    },
    "cmh": {
        "url": "https://aic.sg/Care-Services/Community-Mental-Health-Services",
        "category": "Facility (Healthcare-related)",
        "type": "CMH",
        "description": "Community-based mental health enquiry support, therapy or counselling.",
        "provider": "AIC",
    },
    "mhgp_poly": {
        "url": "https://aic.sg/Care-Services/Mental-Health-GP-Partnership-Polyclinic",
        "category": "Facility (Healthcare-related)",
        "type": "Polyclinic",
        "description": "Polyclinic partner for mental health GP partnership support.",
        "provider": "AIC",
    },
}

SERVICE_ROWS = [
    {
        "[DROPDOWN] Category": "Services",
        "[DROPDOWN] Type": "AAC Services",
        "[FREE TEXT] Name": "Active Ageing Centre Services",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "Active ageing support for seniors, including activities, social engagement and referrals."
        ),
        "[DROPDOWN] Cost": "",
        "[FREE TEXT] Service Provider": "AIC",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://aic.sg/Care-Services/Active-Ageing-Centres",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Services",
        "[DROPDOWN] Type": "SCC Services",
        "[FREE TEXT] Name": "Senior Day Care Services",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "Senior day care support, caregiver support activities and selected dementia day care options."
        ),
        "[DROPDOWN] Cost": "Paid",
        "[FREE TEXT] Service Provider": "AIC",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://aic.sg/Care-Services/Day-Care",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Services",
        "[DROPDOWN] Type": "ServiceSG Services",
        "[FREE TEXT] Name": "ServiceSG Assisted Government Services",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "Walk-in guidance for close to 600 government services, with support for in-person or video appointments."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "ServiceSG / PSD",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.psd.gov.sg/servicesg/",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
]

PROMOTION_ROWS = [
    {
        "[DROPDOWN] Category": "Promotions",
        "[DROPDOWN] Type": "CHAS medical benefits?",
        "[FREE TEXT] Name": "Community Health Assist Scheme (CHAS)",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "Subsidies for medical and dental care at participating GP and dental clinics for Singapore Citizens, including PG and MG cardholders."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "MOH / CHAS",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.chas.sg/about-the-scheme",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Promotions",
        "[DROPDOWN] Type": "PG supermarket benefits",
        "[FREE TEXT] Name": "FairPrice Pioneer Generation Discount Scheme",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "FairPrice discount scheme for Pioneer Generation cardholders on designated shopping days."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "FairPrice",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "Designated shopping days",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.fairprice.com.sg/tips/money-saving/enjoy-greater-value-and-savings-with-fairprice",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Promotions",
        "[DROPDOWN] Type": "MG supermarket benefits",
        "[FREE TEXT] Name": "FairPrice Merdeka Generation Discount Scheme",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "FairPrice discount scheme for Merdeka Generation cardholders on designated shopping days."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "FairPrice",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "Designated shopping days",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.fairprice.com.sg/tips/money-saving/enjoy-greater-value-and-savings-with-fairprice",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Promotions",
        "[DROPDOWN] Type": "CHAS supermarket benefits?",
        "[FREE TEXT] Name": "FairPrice CHAS Blue / Orange Discount Scheme",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "FairPrice discount scheme for CHAS Blue and CHAS Orange cardholders on designated shopping days."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "FairPrice",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "Designated shopping days",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.fairprice.com.sg/tips/money-saving/enjoy-greater-value-and-savings-with-fairprice",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
    {
        "[DROPDOWN] Category": "Promotions",
        "[DROPDOWN] Type": "Corporate promotions",
        "[FREE TEXT] Name": "PAssion Card Member Perks",
        "[FREE TEXT] Description\n(incl. access to which Service or Programme)": (
            "Exclusive perks at Community Clubs, PAssion Wave, imPAct@Hong Lim Green and participating merchants."
        ),
        "[DROPDOWN] Cost": "Free",
        "[FREE TEXT] Service Provider": "People's Association",
        "[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100": "",
        "[FREE TEXT] Operating/Programme Days e.g. Mon-Fri": "",
        "[FREE TEXT] Address": "",
        "[FREE TEXT] Postal Code": "",
        "[FREE TEXT] Contact Number": "",
        "[FREE TEXT] Email": "",
        "[FREE TEXT] Website/URL": "https://www.onepa.gov.sg/passion-card",
        "Lat": "",
        "Long": "",
        "[DROPDOWN] Status": "Active",
        "Planning Area": "",
        "Subzone": "",
        "Constituency": "",
        "RHS": "",
        "Region": "",
        "[DROPDOWN] Sub-region": "",
    },
]

HEB_TEMPLE_URLS = {
    "Sri Mariamman Temple": "https://heb.org.sg/smt/",
    "Sri Srinivasa Perumal Temple": "https://heb.org.sg/sspt/",
    "Sri Sivan Temple": "https://heb.org.sg/sst/",
    "Sri Vairavimada Kaliamman Temple": "https://heb.org.sg/svkt/",
}

ONEPA_DIRECT_OUTLET_URLS = [
    "https://www.onepa.gov.sg/cc/sembawang-cc",
    "https://www.onepa.gov.sg/cc/canberra-cc",
    "https://www.onepa.gov.sg/cc/fernvale-cc",
    "https://www.onepa.gov.sg/rc/buona-vista-zone-d-rn",
    "https://www.onepa.gov.sg/rc/punggol-riviera-rn",
    "https://www.onepa.gov.sg/rc/canberra-northwoods-rn",
    "https://www.onepa.gov.sg/rc/canberra-sembawang-gardens-rn",
    "https://www.onepa.gov.sg/rc/canberra-zone-10-rn",
    "https://www.onepa.gov.sg/rc/canberra-zone-9-rn",
    "https://www.onepa.gov.sg/rc/compassvale-mast-rn",
    "https://www.onepa.gov.sg/rc/woodlands-zone-6-rn",
]

PA_EVENT_URLS = [
    "https://www.onepa.gov.sg/events/puthandu-feast-19959472",
    "https://www.onepa.gov.sg/events/jalan-kayu-csn-coney-island-walk-on-29-mar-2026-95906735",
    "https://www.onepa.gov.sg/events/limbang-ccc-cmwn-rummy-o-3-2026-98956310",
    "https://www.onepa.gov.sg/events/besm-buka-puasa-hari-raya-light-up-2026-54110780",
    "https://www.onepa.gov.sg/events/bbz11-rn-canvas-painting-workshop-60047573",
    "https://www.onepa.gov.sg/events/healthier-sg-hkn-csn-weekly-funwalk-15-march-2026-40515148",
    "https://www.onepa.gov.sg/events/nse-cc-wec-vietnamese-spring-rolls-cooking-demo-28-march-2026-73372168",
    "https://www.onepa.gov.sg/events/pvrn-buka-puasa-2026-60883101",
    "https://www.onepa.gov.sg/events/yuhua-cc-iaec-cooking-demonstration-interest-group-prawn-sambal-with-pilau-rice-56924025",
    "https://www.onepa.gov.sg/events/punggol-waterfront-rn-paint-with-us-at-the-waterfront-community-mural-painting-14-to-17-march-2026-78036213",
]

PA_COURSE_URLS = [
    "https://www.onepa.gov.sg/courses/conversational-japanese-for-beginners-c027195384",
    "https://www.onepa.gov.sg/courses/taekwondo-singapore-taekwondo-federation-healthiersg-c027201712",
    "https://www.onepa.gov.sg/courses/taekwondo-singapore-taekwondo-federation-healthiersg-c027200434",
    "https://www.onepa.gov.sg/courses/hanyu-pinyin-for-k2-c027203757",
    "https://www.onepa.gov.sg/courses/fun-with-keyboard-beginner-c027200530",
    "https://www.onepa.gov.sg/courses/mixed-media-art-c027204136",
    "https://www.onepa.gov.sg/courses/taijiquan-healthiersg-c027203117",
    "https://www.onepa.gov.sg/courses/taekwondo-singapore-taekwondo-federation-healthiersg-c027205200",
    "https://www.onepa.gov.sg/courses/coding-and-robotics-c027208345",
    "https://www.onepa.gov.sg/courses/pa-kiddies-learn2code-roblox-c027205362",
]


def request(method: str, url: str, *, timeout: int = 10, expect_json: bool = False, **kwargs: Any) -> Any:
    kwargs.setdefault("headers", HTTP_HEADERS)
    last_error: Exception | None = None
    for attempt in range(2):
        try:
            response = requests.request(method, url, timeout=timeout, **kwargs)
            if response.status_code in {429, 500, 502, 503, 504}:
                time.sleep((attempt + 1) * 1.5 + random.random())
                continue
            response.raise_for_status()
            if expect_json:
                return response.json()
            return response.text
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep((attempt + 1) * 1.0 + random.random())
    raise RuntimeError(f"Failed request for {url}") from last_error


def clean_html(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html_lib.unescape(text)
    if re.fullmatch(r"https?://\S+", text.strip()):
        return text.strip()
    text = text.replace("<br />", "\n").replace("<br/>", "\n").replace("<br>", "\n")
    text = BeautifulSoup(text, "html.parser").get_text("\n")
    text = text.replace("\xa0", " ")
    lines = [re.sub(r"\s+", " ", line).strip(" ,;") for line in text.splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(lines).strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_html(value)).strip()


def extract_postal(value: str) -> str:
    match = re.search(r"(?<!\d)(\d{6})(?!\d)", value or "")
    return match.group(1) if match else ""


def normalize_phone(value: Any) -> str:
    text = one_line(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def parse_float(value: Any) -> float | str:
    if value in {None, ""}:
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def parse_date(value: str) -> dt.date | None:
    if not value:
        return None
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return dt.datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


def unique_join(values: Iterable[str], sep: str = "; ") -> str:
    seen: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
    return sep.join(seen)


def row_template() -> dict[str, Any]:
    return {header: "" for header in TEMPLATE_HEADERS}


def normalize_address(address: str, postal: str = "") -> str:
    text = address.replace(" ,", ",").replace(" , ", ", ")
    text = re.sub(r"\s+,", ",", text)
    text = re.sub(r"\s+", " ", text).strip(" ,")
    if postal and postal not in text:
        if text:
            text = f"{text}, Singapore {postal}"
        else:
            text = f"Singapore {postal}"
    return text.strip(" ,")


def extract_meta_description(html_text: str) -> str:
    match = re.search(r'<meta name="description" content="([^"]*)"', html_text, re.I)
    return clean_html(match.group(1)) if match else ""


def parse_window_components(html_text: str) -> list[dict[str, Any]]:
    components: list[dict[str, Any]] = []
    for match in re.finditer(r"window\.reactComponents\.push\((\{.*?\})\);</script>", html_text, re.S):
        try:
            components.append(json.loads(match.group(1)))
        except json.JSONDecodeError:
            continue
    return components


def item_fields(item_fields_list: list[dict[str, Any]]) -> dict[str, Any]:
    return {field["name"]: field.get("jsonValue") for field in item_fields_list}


def request_data_gov_download(dataset_id: str) -> dict[str, Any]:
    poll = request(
        "GET",
        f"https://api-open.data.gov.sg/v1/public/api/datasets/{dataset_id}/poll-download",
        expect_json=True,
    )
    return request("GET", poll["data"]["url"], timeout=90, expect_json=True)


def point_in_ring(point_x: float, point_y: float, ring: list[list[float]]) -> bool:
    inside = False
    for i in range(len(ring)):
        x1, y1 = ring[i]
        x2, y2 = ring[(i + 1) % len(ring)]
        if ((y1 > point_y) != (y2 > point_y)) and (
            point_x < (x2 - x1) * (point_y - y1) / ((y2 - y1) or 1e-12) + x1
        ):
            inside = not inside
    return inside


def point_in_polygon(point_x: float, point_y: float, polygon: list[list[list[float]]]) -> bool:
    if not polygon:
        return False
    if not point_in_ring(point_x, point_y, polygon[0]):
        return False
    for hole in polygon[1:]:
        if point_in_ring(point_x, point_y, hole):
            return False
    return True


def prepare_geo_lookup(dataset_id: str) -> list[dict[str, Any]]:
    data = request_data_gov_download(dataset_id)
    prepared: list[dict[str, Any]] = []
    for feature in data["features"]:
        geometry = feature["geometry"]
        polygons: list[list[list[list[float]]]] = []
        if geometry["type"] == "Polygon":
            polygons = [geometry["coordinates"]]
        elif geometry["type"] == "MultiPolygon":
            polygons = geometry["coordinates"]
        else:
            continue
        xs: list[float] = []
        ys: list[float] = []
        for polygon in polygons:
            for ring in polygon:
                for lon, lat in ring:
                    xs.append(lon)
                    ys.append(lat)
        prepared.append(
            {
                "properties": feature["properties"],
                "polygons": polygons,
                "bbox": (min(xs), min(ys), max(xs), max(ys)),
            }
        )
    return prepared


def locate_feature(lat: float, lon: float, prepared_features: list[dict[str, Any]]) -> dict[str, Any] | None:
    for feature in prepared_features:
        min_x, min_y, max_x, max_y = feature["bbox"]
        if lon < min_x or lon > max_x or lat < min_y or lat > max_y:
            continue
        for polygon in feature["polygons"]:
            if point_in_polygon(lon, lat, polygon):
                return feature["properties"]
    return None


def split_schedule(*values: str) -> tuple[str, str]:
    combined = "\n".join([clean_html(value) for value in values if clean_html(value)])
    if not combined:
        return "", ""
    days: list[str] = []
    hours: list[str] = []
    for line in combined.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(
            r"^(Mon(?:day)?|Tue(?:sday)?|Wed(?:nesday)?|Thu(?:rsday)?|Fri(?:day)?|Sat(?:urday)?|Sun(?:day)?|Weekdays?|Weekends?)(.*?):\s*(.+)$",
            line,
            re.I,
        )
        if match:
            days.append(f"{match.group(1)}{match.group(2)}".strip())
            hours.append(match.group(3).strip())
            continue
        if re.search(r"\b(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\b", line, re.I):
            days.append(line)
            continue
        hours.append(line)
    if not days and not hours:
        return "", combined
    return unique_join(days), unique_join(hours) or combined


def normalize_cost_from_prices(*values: str | None) -> str:
    filtered = [value for value in values if value]
    if not filtered:
        return ""
    normalized = " ".join(filtered).lower()
    if "free" in normalized or "$0" in normalized:
        return "Free"
    return "Paid"


def validate_postal(postal: str, valid_postals: set[str]) -> str:
    if not postal:
        return ""
    if re.fullmatch(r"\d{6}", postal):
        return postal
    return ""


def load_valid_postals() -> set[str]:
    wb = load_workbook(POSTAL_CODES_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    valid: set[str] = set()
    for cell, in ws.iter_rows(min_row=2, values_only=True):
        if cell:
            valid.add(str(cell).zfill(6))
    return valid


def find_aic_page_context(page_url: str) -> dict[str, Any]:
    html_text = request("GET", page_url)
    next_data_match = re.search(
        r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>',
        html_text,
        re.S,
    )
    if not next_data_match:
        raise RuntimeError(f"Unable to parse AIC page context for {page_url}")
    data = json.loads(next_data_match.group(1))
    route = data["props"]["pageProps"]["page"]["layout"]["sitecore"]["route"]

    map_component: dict[str, Any] | None = None

    def walk(node: Any) -> None:
        nonlocal map_component
        if map_component:
            return
        if isinstance(node, dict):
            if node.get("componentName") == "MapSelectorWithFilters":
                map_component = node
                return
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for value in node:
                walk(value)

    walk(route)
    if not map_component:
        raise RuntimeError(f"MapSelectorWithFilters not found for {page_url}")
    datasource_fields = item_fields(map_component["fields"]["data"]["datasource"]["fields"])
    return {
        "selected_categories": datasource_fields.get("SelectedCategory", []),
        "short_description": clean_html(route["fields"].get("ShortDescription", {}).get("value", "")),
        "meta_description": extract_meta_description(html_text),
    }


def fetch_aic_map_items(category_ids: list[str]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    cursor: str | None = None
    while True:
        payload: dict[str, Any] = {"language": "en", "first": 100, "categoryIds": category_ids}
        if cursor:
            payload["after"] = cursor
        response = request(
            "POST",
            "https://www.aic.sg/api/map-items",
            json=payload,
            expect_json=True,
        )
        data = response["data"]
        items.extend(data["items"])
        if not data["hasNextPage"]:
            break
        cursor = data["endCursor"]
    return items


def build_aic_rows(valid_postals: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    contexts: dict[str, dict[str, Any]] = {
        key: find_aic_page_context(spec["url"]) for key, spec in AIC_PAGES.items()
    }
    for key, spec in AIC_PAGES.items():
        category_ids = [item["id"] for item in contexts[key]["selected_categories"]]
        items = fetch_aic_map_items(category_ids)
        for item in items:
            fields = item_fields(item["fields"])
            title = one_line(fields.get("Title", {}).get("value", ""))
            if key == "mhgp_poly" and "polyclinic" not in title.lower():
                continue
            address_text = clean_html(fields.get("Address", {}).get("value", ""))
            postal = validate_postal(extract_postal(address_text), valid_postals)
            description_parts = [spec["description"]]
            subcategories = [entry.get("displayName", "") for entry in (fields.get("Subcategory") or [])]
            if key == "day_care" and subcategories:
                description_parts.append(unique_join(subcategories))
            if key == "nursing_home" and subcategories:
                description_parts.append(unique_join(subcategories))
            if key == "cmh" and subcategories:
                description_parts.append(unique_join(subcategories))
            notes = one_line(fields.get("Notes", {}).get("value", ""))
            if notes:
                description_parts.append(notes)

            row = row_template()
            row["[DROPDOWN] Category"] = spec["category"]
            row["[DROPDOWN] Type"] = spec["type"]
            row["[FREE TEXT] Name"] = title
            row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = unique_join(
                description_parts
            )
            if key == "day_care":
                row["[DROPDOWN] Cost"] = "Paid"
            row["[FREE TEXT] Service Provider"] = spec["provider"]
            row["[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100"] = one_line(
                fields.get("OperatingHours", {}).get("value", "")
            )
            row["[FREE TEXT] Operating/Programme Days e.g. Mon-Fri"] = ""
            row["[FREE TEXT] Address"] = normalize_address(one_line(address_text), postal)
            row["[FREE TEXT] Postal Code"] = postal
            row["[FREE TEXT] Contact Number"] = normalize_phone(fields.get("ContactNumber", {}).get("value", ""))
            row["[FREE TEXT] Email"] = one_line(fields.get("Email", {}).get("value", ""))
            row["[FREE TEXT] Website/URL"] = one_line(fields.get("WebsiteURL", {}).get("value", "")) or spec["url"]
            row["Lat"] = parse_float(fields.get("Latitude", {}).get("value", ""))
            row["Long"] = parse_float(fields.get("Longitude", {}).get("value", ""))
            row["[DROPDOWN] Status"] = "Active"
            rows.append(row)
    return rows


def parse_onepa_outlet_page(url: str, valid_postals: set[str]) -> dict[str, Any] | None:
    html_text = request("GET", url)
    components = parse_window_components(html_text)
    detail_data: dict[str, Any] | None = None
    location_data: dict[str, Any] | None = None
    for component in components:
        data = component.get("data")
        if component.get("component") == "CCDetailInformation" and isinstance(data, dict):
            detail_data = data
        if component.get("component") == "CcDetailsSingleLocation" and isinstance(data, dict):
            location_data = data
    if not detail_data:
        return None

    content = detail_data.get("content", {})
    address_obj = content.get("address", {})
    address_desc = clean_html(address_obj.get("description", ""))
    postal = validate_postal(str(address_obj.get("pinCode", "")).strip(), valid_postals)
    full_address = normalize_address(address_desc, postal)
    title = one_line(detail_data.get("name", ""))
    path = urlparse(url).path

    if path.startswith("/cc/"):
        facility_type = "CC (Temp Office)" if "currently operating from" in address_desc.lower() else "CC"
    elif re.search(r"\bRN\b", title):
        facility_type = "RN (Residents' Network)"
    elif re.search(r"\bNC\b", title):
        facility_type = "NC (Neighbourhood Committee)"
    else:
        facility_type = "RC (Residents' Corner)"

    days, hours = split_schedule(
        content.get("officeHours", {}).get("openingDays", ""),
        content.get("officeHours", {}).get("openingHours", ""),
    )

    row = row_template()
    row["[DROPDOWN] Category"] = "Facility (Govt Community-related)"
    row["[DROPDOWN] Type"] = facility_type
    row["[FREE TEXT] Name"] = title
    row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = one_line(
        content.get("subHeader", "")
    )
    row["[FREE TEXT] Service Provider"] = "People's Association"
    row["[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100"] = hours
    row["[FREE TEXT] Operating/Programme Days e.g. Mon-Fri"] = days
    row["[FREE TEXT] Address"] = full_address
    row["[FREE TEXT] Postal Code"] = postal
    row["[FREE TEXT] Contact Number"] = normalize_phone(content.get("contact", {}).get("number", ""))
    row["[FREE TEXT] Email"] = one_line(content.get("contact", {}).get("email", ""))
    row["[FREE TEXT] Website/URL"] = url
    if location_data:
        row["Lat"] = parse_float(location_data.get("latitude", ""))
        row["Long"] = parse_float(location_data.get("longitude", ""))
    row["[DROPDOWN] Status"] = "Active"
    row["_cache_url"] = url
    row["_cache_path"] = path
    return row


def fetch_onepa_outlets(valid_postals: set[str]) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    cache: dict[str, dict[str, Any]] = {}

    def worker(url: str) -> dict[str, Any] | None:
        try:
            return parse_onepa_outlet_page(url, valid_postals)
        except Exception:  # noqa: BLE001
            return None

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for result in executor.map(worker, ONEPA_DIRECT_OUTLET_URLS):
            if not result:
                continue
            cache_key_url = result.pop("_cache_url")
            cache_key_path = result.pop("_cache_path")
            rows.append(result)
            cache[cache_key_url] = copy.deepcopy(result)
            cache[cache_key_path] = copy.deepcopy(result)

    dedupe: set[tuple[str, str]] = {(row["[FREE TEXT] Name"], row["[FREE TEXT] Postal Code"]) for row in rows}
    community_rows = onemap_search_all("community club") + onemap_search_all("community centre")
    for item in community_rows:
        name = one_line(item.get("SEARCHVAL", ""))
        if not ("community club" in name.lower() or "community centre" in name.lower()):
            continue
        address = one_line(item.get("ADDRESS", ""))
        postal = validate_postal(str(item.get("POSTAL", "")).strip(), valid_postals)
        key = (name, postal or address)
        if key in dedupe:
            continue
        dedupe.add(key)
        row = row_template()
        row["[DROPDOWN] Category"] = "Facility (Govt Community-related)"
        row["[DROPDOWN] Type"] = "CC"
        row["[FREE TEXT] Name"] = name
        row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = (
            "People's Association community club or community centre."
        )
        row["[FREE TEXT] Service Provider"] = "People's Association"
        row["[FREE TEXT] Address"] = normalize_address(address, postal)
        row["[FREE TEXT] Postal Code"] = postal
        row["[FREE TEXT] Website/URL"] = (
            f"https://www.onemap.gov.sg/?{urlencode({'lat': item.get('LATITUDE', ''), 'lng': item.get('LONGITUDE', '')})}"
        )
        row["Lat"] = parse_float(item.get("LATITUDE", ""))
        row["Long"] = parse_float(item.get("LONGITUDE", ""))
        row["[DROPDOWN] Status"] = "Active"
        rows.append(row)
    return rows, cache


def parse_query_lat_lon(url: str) -> tuple[float | str, float | str]:
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    lat = query.get("lat", [""])[0]
    lng = query.get("lng", [""])[0]
    return parse_float(lat), parse_float(lng)


def split_servicesg_hours(text_values: list[str]) -> tuple[str, str]:
    if not text_values:
        return "", ""
    joined = " ".join(text_values)
    match = re.search(r"Open\s+(.+?)\s+daily", joined, re.I)
    hours = match.group(1).strip() if match else unique_join(text_values)
    if "excluding Sundays and public holidays" in joined:
        days = "Daily (excluding Sundays and public holidays)"
    elif "excluding public holidays" in joined:
        days = "Daily (excluding public holidays)"
    else:
        days = "Daily"
    return days, hours


def fetch_servicesg_centres(valid_postals: set[str]) -> list[dict[str, Any]]:
    html_text = request("GET", "https://www.life.gov.sg/services/book-virtual-appointment-servicesg")
    soup = BeautifulSoup(html_text, "html.parser")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for anchor in soup.find_all("a", href=re.compile(r"onemap\.gov\.sg/\?lat=")):
        title_tag = anchor.find(["h6", "h5"])
        if not title_tag:
            continue
        name = one_line(title_tag.get_text(" ", strip=True))
        address_lines = [one_line(tag.get_text(" ", strip=True)) for tag in anchor.find_all("p")]
        address = address_lines[0] if address_lines else ""
        if not address or address in seen:
            continue
        seen.add(address)
        postal = validate_postal(extract_postal(address), valid_postals)
        days, hours = split_servicesg_hours(address_lines[1:])
        lat, lng = parse_query_lat_lon(anchor.get("href", ""))

        row = row_template()
        row["[DROPDOWN] Category"] = "Facility (Govt Community-related)"
        row["[DROPDOWN] Type"] = "ServiceSG Centre"
        row["[FREE TEXT] Name"] = name if name.lower().startswith("servicesg") else f"ServiceSG Centre {name}"
        row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = (
            "Walk-in support for government services and appointments."
        )
        row["[DROPDOWN] Cost"] = "Free"
        row["[FREE TEXT] Service Provider"] = "ServiceSG / PSD"
        row["[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100"] = hours
        row["[FREE TEXT] Operating/Programme Days e.g. Mon-Fri"] = days
        row["[FREE TEXT] Address"] = normalize_address(address, postal)
        row["[FREE TEXT] Postal Code"] = postal
        row["[FREE TEXT] Website/URL"] = "https://www.life.gov.sg/services/book-virtual-appointment-servicesg"
        row["Lat"] = lat
        row["Long"] = lng
        row["[DROPDOWN] Status"] = "Active"
        rows.append(row)
    return rows


def onemap_search(query: str, page_num: int = 1) -> dict[str, Any]:
    params = {
        "searchVal": query,
        "returnGeom": "Y",
        "getAddrDetails": "Y",
        "pageNum": page_num,
    }
    return request(
        "GET",
        "https://www.onemap.gov.sg/api/common/elastic/search",
        params=params,
        expect_json=True,
    )


def onemap_search_all(query: str) -> list[dict[str, Any]]:
    first_page = onemap_search(query, 1)
    results = list(first_page.get("results", []))
    total_pages = int(first_page.get("totalNumPages") or 1)
    for page_num in range(2, total_pages + 1):
        time.sleep(0.6)
        page = onemap_search(query, page_num)
        results.extend(page.get("results", []))
    return results


def build_onemap_religious_row(
    name: str,
    facility_type: str,
    address: str,
    postal: str,
    latitude: str,
    longitude: str,
    source_url: str,
) -> dict[str, Any]:
    row = row_template()
    row["[DROPDOWN] Category"] = "Facility (Religious-related)"
    row["[DROPDOWN] Type"] = facility_type
    row["[FREE TEXT] Name"] = name
    row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = (
        "Religious organisation / place of worship."
    )
    row["[FREE TEXT] Address"] = normalize_address(address, postal)
    row["[FREE TEXT] Postal Code"] = postal
    row["[FREE TEXT] Website/URL"] = source_url
    row["Lat"] = parse_float(latitude)
    row["Long"] = parse_float(longitude)
    row["[DROPDOWN] Status"] = "Active"
    return row


def fetch_religious_organisations(valid_postals: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()

    query_specs = [
        ("masjid", "Mosque", lambda n: ("MASJID" in n or "MOSQUE" in n) and n != "MOSQUE"),
        (
            "church",
            "Church",
            lambda n: any(token in n for token in ["CHURCH", "CATHEDRAL", "CHAPEL"]),
        ),
        (
            "chinese temple",
            "Chinese Temple",
            lambda n: "TEMPLE" in n and n not in {"CHINESE TEMPLE"},
        ),
        (
            "gurdwara",
            "Sikh Temple",
            lambda n: "GURDWARA" in n and "HISTORIC SITE" not in n,
        ),
        (
            "synagogue",
            "Synagogue",
            lambda n: "SYNAGOGUE" in n and "STREET" not in n,
        ),
    ]

    for query, facility_type, validator in query_specs:
        for item in onemap_search_all(query):
            name = one_line(item.get("SEARCHVAL", ""))
            if not validator(name.upper()):
                continue
            address = one_line(item.get("ADDRESS", ""))
            postal = validate_postal(str(item.get("POSTAL", "")).strip(), valid_postals)
            if not postal:
                postal = validate_postal(extract_postal(address), valid_postals)
            dedupe_key = (facility_type, name, postal or address)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            source_url = f"https://www.onemap.gov.sg/?{urlencode({'lat': item.get('LATITUDE', ''), 'lng': item.get('LONGITUDE', '')})}"
            rows.append(
                build_onemap_religious_row(
                    name=name,
                    facility_type=facility_type,
                    address=address,
                    postal=postal,
                    latitude=item.get("LATITUDE", ""),
                    longitude=item.get("LONGITUDE", ""),
                    source_url=source_url,
                )
            )

    for temple_name, temple_url in HEB_TEMPLE_URLS.items():
        results = onemap_search_all(temple_name)
        if not results:
            continue
        item = results[0]
        postal = validate_postal(str(item.get("POSTAL", "")).strip(), valid_postals)
        rows.append(
            build_onemap_religious_row(
                name=temple_name,
                facility_type="Indian Temple",
                address=one_line(item.get("ADDRESS", "")),
                postal=postal,
                latitude=item.get("LATITUDE", ""),
                longitude=item.get("LONGITUDE", ""),
                source_url=temple_url,
            )
        )
    return rows


def parse_onepa_commerce_page(url: str, component_name: str) -> dict[str, Any] | None:
    html_text = request("GET", url)
    for component in parse_window_components(html_text):
        if component.get("component") == component_name:
            return component.get("data")
    return None


def fetch_pa_programmes_events(
    valid_postals: set[str],
    outlet_cache: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for url in PA_EVENT_URLS:
        try:
            data = parse_onepa_commerce_page(url, "EventDetails")
        except Exception:  # noqa: BLE001
            continue
        if not data:
            continue
        venue = data.get("venue", {})
        outlet_info = outlet_cache.get(urljoin("https://www.onepa.gov.sg", data.get("outletUrl", "")), {})
        event_date = parse_date(data.get("sessions", {}).get("date", ""))
        postal = validate_postal(venue.get("addressPostalCode", ""), valid_postals)
        if not postal:
            postal = outlet_info.get("[FREE TEXT] Postal Code", "")
        address = one_line(venue.get("address", "")) or outlet_info.get("[FREE TEXT] Address", "")

        row = row_template()
        row["[DROPDOWN] Category"] = "Programmes/Events"
        row["[DROPDOWN] Type"] = "PA Programme/Events"
        row["[FREE TEXT] Name"] = one_line(data.get("heading", ""))
        row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = one_line(
            data.get("description", "")
        )
        row["[DROPDOWN] Cost"] = normalize_cost_from_prices(
            data.get("price", {}).get("memberPrice"), data.get("price", {}).get("nonMemberPrice")
        )
        row["[FREE TEXT] Service Provider"] = one_line(
            data.get("mainOrganisingCommitteeName") or data.get("organiser", {}).get("name") or "PA"
        )
        row["[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100"] = one_line(
            data.get("sessions", {}).get("time", "")
        )
        row["[FREE TEXT] Operating/Programme Days e.g. Mon-Fri"] = one_line(
            data.get("sessions", {}).get("date", "")
        )
        row["[FREE TEXT] Address"] = normalize_address(address, postal)
        row["[FREE TEXT] Postal Code"] = postal
        row["[FREE TEXT] Website/URL"] = url
        row["Lat"] = parse_float(venue.get("mapLat", ""))
        row["Long"] = parse_float(venue.get("mapLon", ""))
        row["[DROPDOWN] Status"] = "Active" if (event_date and event_date >= TODAY) else "Inactive"
        rows.append(row)

    for url in PA_COURSE_URLS:
        try:
            data = parse_onepa_commerce_page(url, "CourseDetails")
        except Exception:  # noqa: BLE001
            continue
        if not data:
            continue
        outlet_url = urljoin("https://www.onepa.gov.sg", data.get("outletUrl", ""))
        outlet_info = outlet_cache.get(outlet_url, {})
        end_date = parse_date(data.get("sessions", {}).get("endDate", ""))

        row = row_template()
        row["[DROPDOWN] Category"] = "Programmes/Events"
        row["[DROPDOWN] Type"] = "PA Programme/Events"
        row["[FREE TEXT] Name"] = one_line(data.get("heading", ""))
        row["[FREE TEXT] Description\n(incl. access to which Service or Programme)"] = one_line(
            data.get("description", "")
        )
        row["[DROPDOWN] Cost"] = normalize_cost_from_prices(
            data.get("price", {}).get("memberPrice"),
            data.get("price", {}).get("nonMemberPrice"),
        )
        row["[FREE TEXT] Service Provider"] = one_line(
            data.get("mainOrganisingCommitteeName") or data.get("outletName") or "PA"
        )
        session = data.get("sessions", {})
        row["[FREE TEXT] Operating/Programme Hours (24h) e.g. 0900-1100"] = unique_join(
            [one_line(session.get("startTime", "")), one_line(session.get("endTime", ""))],
            sep=" to ",
        )
        row["[FREE TEXT] Operating/Programme Days e.g. Mon-Fri"] = unique_join(
            [one_line(session.get("day", "")), one_line(session.get("startDate", "")), one_line(session.get("endDate", ""))]
        )
        row["[FREE TEXT] Address"] = outlet_info.get("[FREE TEXT] Address", "")
        row["[FREE TEXT] Postal Code"] = outlet_info.get("[FREE TEXT] Postal Code", "")
        row["[FREE TEXT] Contact Number"] = outlet_info.get("[FREE TEXT] Contact Number", "")
        row["[FREE TEXT] Email"] = outlet_info.get("[FREE TEXT] Email", "")
        row["[FREE TEXT] Website/URL"] = url
        row["Lat"] = parse_float(data.get("map", {}).get("latitude", ""))
        row["Long"] = parse_float(data.get("map", {}).get("longitude", ""))
        row["[DROPDOWN] Status"] = "Active" if (end_date and end_date >= TODAY) else "Inactive"
        rows.append(row)

    return rows


def enrich_geography(
    rows: list[dict[str, Any]],
    subzone_lookup: list[dict[str, Any]],
    constituency_lookup: list[dict[str, Any]],
) -> None:
    for row in rows:
        lat = row.get("Lat")
        lng = row.get("Long")
        if lat == "" or lng == "":
            continue
        if not isinstance(lat, float) or not isinstance(lng, float):
            continue
        subzone = locate_feature(lat, lng, subzone_lookup)
        if subzone:
            row["Planning Area"] = subzone.get("PLN_AREA_N", "")
            row["Subzone"] = subzone.get("SUBZONE_N", "")
            row["Region"] = subzone.get("REGION_N", "")
        constituency = locate_feature(lat, lng, constituency_lookup)
        if constituency:
            row["Constituency"] = constituency.get("ED_DESC_FU", "") or constituency.get("ED_DESC", "")


def sanitize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sanitized: list[dict[str, Any]] = []
    for row in rows:
        clean_row = row_template()
        for header in TEMPLATE_HEADERS:
            clean_row[header] = row.get(header, "")
        sanitized.append(clean_row)
    return sanitized


def write_workbook(rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    if headers != TEMPLATE_HEADERS:
        raise RuntimeError("Template headers do not match expected structure")
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append([row.get(header, "") for header in TEMPLATE_HEADERS])
    wb.save(OUTPUT_PATH)


def main() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    valid_postals = load_valid_postals()
    subzone_lookup = prepare_geo_lookup("d_8594ae9ff96d0c708bc2af633048edfb")
    constituency_lookup = prepare_geo_lookup("d_7ddf956dfc1c59080bf95bba1c58a5d2")
    print("Loaded postal and boundary lookups", flush=True)

    aic_rows = build_aic_rows(valid_postals)
    print(f"AIC rows: {len(aic_rows)}", flush=True)
    onepa_rows, outlet_cache = fetch_onepa_outlets(valid_postals)
    print(f"onePA outlet rows: {len(onepa_rows)}", flush=True)
    servicesg_rows = fetch_servicesg_centres(valid_postals)
    print(f"ServiceSG rows: {len(servicesg_rows)}", flush=True)
    religious_rows = fetch_religious_organisations(valid_postals)
    print(f"Religious rows: {len(religious_rows)}", flush=True)
    pa_programme_rows = fetch_pa_programmes_events(valid_postals, outlet_cache)
    print(f"PA programme/event rows: {len(pa_programme_rows)}", flush=True)

    all_rows = (
        aic_rows
        + onepa_rows
        + servicesg_rows
        + religious_rows
        + SERVICE_ROWS
        + pa_programme_rows
        + PROMOTION_ROWS
    )

    enrich_geography(all_rows, subzone_lookup, constituency_lookup)
    print("Geography enrichment complete", flush=True)
    final_rows = sanitize_rows(all_rows)
    final_rows.sort(
        key=lambda row: (
            row.get("[DROPDOWN] Category", ""),
            row.get("[DROPDOWN] Type", ""),
            row.get("[FREE TEXT] Name", ""),
        )
    )

    write_workbook(final_rows)
    print("Workbook written", flush=True)
    print(json.dumps(
        {
            "output": str(OUTPUT_PATH),
            "row_count": len(final_rows),
            "counts": {
                "aic": len(aic_rows),
                "onepa_outlets": len(onepa_rows),
                "servicesg": len(servicesg_rows),
                "religious": len(religious_rows),
                "pa_programmes_events": len(pa_programme_rows),
                "soft_resources": len(SERVICE_ROWS) + len(PROMOTION_ROWS),
            },
        },
        indent=2,
    ))


if __name__ == "__main__":
    main()
